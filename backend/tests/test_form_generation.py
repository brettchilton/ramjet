"""
Tests for Phase 4: Form generation, approve/reject, update, and download endpoints.

Run inside Docker:
    docker exec eezy_peezy_backend python -m pytest tests/test_form_generation.py -v
"""

import uuid
from decimal import Decimal
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.main import app
from app.core.database import get_db, SessionLocal
from app.core.models import (
    Order, OrderLineItem, IncomingEmail, EmailAttachment,
)
from app.services.form_generation_service import (
    generate_office_order,
    generate_works_order,
    generate_all_forms,
)

client = TestClient(app)


@pytest.fixture
def db():
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


# ── Helpers ───────────────────────────────────────────────────────────

def _create_test_email(db: Session) -> IncomingEmail:
    email = IncomingEmail(
        gmail_message_id=f"test-form-{uuid.uuid4().hex[:8]}",
        sender="customer@example.com",
        subject="PO-TEST-FORM",
        body_text="Test order for form generation.",
        processed=True,
    )
    db.add(email)
    db.commit()
    db.refresh(email)
    return email


def _create_test_order(db: Session, email: IncomingEmail, with_items: bool = True) -> Order:
    order = Order(
        email_id=email.id,
        status="pending",
        customer_name="Test Customer Pty Ltd",
        po_number="PO-TEST-001",
        special_instructions="Handle with care. Deliver to loading dock B.",
    )
    db.add(order)
    db.flush()

    if with_items:
        item1 = OrderLineItem(
            order_id=order.id,
            line_number=1,
            product_code="LOCAP2",
            matched_product_code="LOCAP2",
            product_description="LOUVRE END CAP 152mm",
            colour="Black",
            quantity=1000,
            unit_price=Decimal("1.32"),
            line_total=Decimal("1320.00"),
        )
        item2 = OrderLineItem(
            order_id=order.id,
            line_number=2,
            product_code="UNKNOWNPROD",
            matched_product_code=None,
            product_description="Unknown widget",
            colour="White",
            quantity=500,
            unit_price=Decimal("2.50"),
            line_total=Decimal("1250.00"),
            needs_review=True,
        )
        db.add_all([item1, item2])

    db.commit()
    db.refresh(order)
    return order


def _cleanup(db: Session, email_ids: list[int]):
    db.rollback()  # Reset any aborted transaction state
    for eid in email_ids:
        orders = db.query(Order).filter(Order.email_id == eid).all()
        for order in orders:
            db.query(OrderLineItem).filter(OrderLineItem.order_id == order.id).delete()
            db.delete(order)
        db.flush()  # Flush order deletes before removing emails (FK constraint)
        db.query(EmailAttachment).filter(EmailAttachment.email_id == eid).delete()
        db.query(IncomingEmail).filter(IncomingEmail.id == eid).delete()
    db.commit()


# ── Office Order Generation Tests ─────────────────────────────────────

class TestOfficeOrderGeneration:
    def test_generate_office_order_structure(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)
            xlsx_bytes = generate_office_order(db, order)

            assert xlsx_bytes is not None
            assert len(xlsx_bytes) > 0

            wb = load_workbook(BytesIO(xlsx_bytes))
            ws = wb.active

            # Title rows
            assert ws["A1"].value == "EEZY PEEZY PLASTICS"
            assert ws["A2"].value == "OFFICE ORDER"

            # Customer / PO
            assert ws["A4"].value == "Customer:"
            assert ws["B4"].value == "Test Customer Pty Ltd"
            assert ws["E4"].value == "PO #:"
            assert ws["F4"].value == "PO-TEST-001"

            # Table headers (row 7)
            assert ws.cell(row=7, column=1).value == "Line"
            assert ws.cell(row=7, column=2).value == "Product Code"
            assert ws.cell(row=7, column=7).value == "Line Total"

            # Line items (rows 8-9)
            assert ws.cell(row=8, column=1).value == 1
            assert ws.cell(row=8, column=2).value == "LOCAP2"
            assert ws.cell(row=8, column=5).value == 1000

            assert ws.cell(row=9, column=1).value == 2
            assert ws.cell(row=9, column=2).value == "UNKNOWNPROD"
            assert ws.cell(row=9, column=5).value == 500

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_office_order_totals(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)
            xlsx_bytes = generate_office_order(db, order)
            wb = load_workbook(BytesIO(xlsx_bytes))
            ws = wb.active

            # Total row should be row 11 (2 items at rows 8-9, blank row 10, total row 11)
            # Find the total row
            total_value = None
            for row in range(10, 15):
                cell = ws.cell(row=row, column=1)
                if cell.value and "Order Total" in str(cell.value):
                    total_value = ws.cell(row=row, column=7).value
                    break

            assert total_value is not None
            assert float(total_value) == 2570.00  # 1320.00 + 1250.00

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_office_order_special_instructions(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)
            xlsx_bytes = generate_office_order(db, order)
            wb = load_workbook(BytesIO(xlsx_bytes))
            ws = wb.active

            # Find special instructions
            found = False
            for row in range(1, 20):
                cell = ws.cell(row=row, column=1)
                if cell.value and "Handle with care" in str(cell.value):
                    found = True
                    break
            assert found, "Special instructions not found in spreadsheet"

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_office_order_no_special_instructions(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)
            order.special_instructions = None
            db.commit()
            db.refresh(order)

            xlsx_bytes = generate_office_order(db, order)
            assert xlsx_bytes is not None
            assert len(xlsx_bytes) > 0

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_office_order_missing_fields(self, db):
        email = _create_test_email(db)
        try:
            order = Order(
                email_id=email.id,
                status="pending",
                customer_name=None,
                po_number=None,
            )
            db.add(order)
            db.flush()

            item = OrderLineItem(
                order_id=order.id,
                line_number=1,
                product_code=None,
                quantity=100,
                unit_price=None,
                line_total=None,
            )
            db.add(item)
            db.commit()
            db.refresh(order)

            # Should not raise even with missing fields
            xlsx_bytes = generate_office_order(db, order)
            assert xlsx_bytes is not None

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise


# ── Works Order Generation Tests ──────────────────────────────────────

class TestWorksOrderGeneration:
    def test_matched_product_works_order(self, db):
        """Works order for a matched product should have populated spec sections."""
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)
            matched_item = next(li for li in order.line_items if li.matched_product_code == "LOCAP2")

            xlsx_bytes = generate_works_order(db, order, matched_item)
            assert xlsx_bytes is not None

            wb = load_workbook(BytesIO(xlsx_bytes))
            ws = wb.active

            # Title
            assert ws["A1"].value == "EEZY PEEZY PLASTICS"
            assert ws["A2"].value == "WORKS ORDER"
            assert ws["A3"].value == "WO-PO-TEST-001-1"

            # Check that Manufacturing Specifications section exists
            found_mfg = False
            for row in range(1, 50):
                cell = ws.cell(row=row, column=1)
                if cell.value and "Manufacturing Specifications" in str(cell.value):
                    found_mfg = True
                    break
            assert found_mfg, "Manufacturing Specifications section not found"

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_unmatched_product_works_order(self, db):
        """Works order for an unmatched product should show unavailable messages."""
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)
            unmatched_item = next(li for li in order.line_items if li.matched_product_code is None)

            xlsx_bytes = generate_works_order(db, order, unmatched_item)
            assert xlsx_bytes is not None

            wb = load_workbook(BytesIO(xlsx_bytes))
            ws = wb.active

            assert ws["A3"].value == "WO-PO-TEST-001-2"

            # Check for "unavailable" message
            found_unavailable = False
            for row in range(1, 50):
                cell = ws.cell(row=row, column=1)
                if cell.value and "specifications unavailable" in str(cell.value):
                    found_unavailable = True
                    break
            assert found_unavailable, "Expected 'unavailable' message for unmatched product"

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_wo_number_format(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)
            item = order.line_items[0]

            xlsx_bytes = generate_works_order(db, order, item)
            wb = load_workbook(BytesIO(xlsx_bytes))
            ws = wb.active

            # WO# should be in row 3
            assert ws["A3"].value.startswith("WO-")
            assert "PO-TEST-001" in ws["A3"].value

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise


# ── Generate All Forms Tests ──────────────────────────────────────────

class TestGenerateAllForms:
    def test_generate_all_forms(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)

            assert order.office_order_file is None
            for item in order.line_items:
                assert item.works_order_file is None

            generate_all_forms(db, order)

            db.refresh(order)
            assert order.office_order_file is not None
            assert len(order.office_order_file) > 0

            for item in order.line_items:
                db.refresh(item)
                assert item.works_order_file is not None
                assert len(item.works_order_file) > 0

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise


# ── Approve Workflow Tests ────────────────────────────────────────────

class TestApproveWorkflow:
    def test_approve_pending_order(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)

            resp = client.post(f"/api/orders/{order.id}/approve")
            assert resp.status_code == 200

            data = resp.json()
            assert data["status"] == "approved"
            assert data["office_order_generated"] is True
            assert data["works_orders_generated"] == 2

            # Verify in DB
            db.refresh(order)
            assert order.status == "approved"
            assert order.office_order_file is not None
            assert order.approved_at is not None

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_approve_already_approved(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)
            order.status = "approved"
            db.commit()

            resp = client.post(f"/api/orders/{order.id}/approve")
            assert resp.status_code == 400
            assert "pending" in resp.json()["detail"]

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_approve_no_line_items(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email, with_items=False)

            resp = client.post(f"/api/orders/{order.id}/approve")
            assert resp.status_code == 400
            assert "no line items" in resp.json()["detail"]

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_approve_nonexistent_order(self):
        fake_id = str(uuid.uuid4())
        resp = client.post(f"/api/orders/{fake_id}/approve")
        assert resp.status_code == 404


# ── Reject Workflow Tests ─────────────────────────────────────────────

class TestRejectWorkflow:
    def test_reject_pending_order(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)

            resp = client.post(
                f"/api/orders/{order.id}/reject",
                json={"reason": "Incorrect pricing"},
            )
            assert resp.status_code == 200

            data = resp.json()
            assert data["status"] == "rejected"
            assert data["message"] == "Order rejected"

            db.refresh(order)
            assert order.status == "rejected"
            assert order.rejected_reason == "Incorrect pricing"

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_reject_already_approved(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)
            order.status = "approved"
            db.commit()

            resp = client.post(
                f"/api/orders/{order.id}/reject",
                json={"reason": "Changed mind"},
            )
            assert resp.status_code == 400

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_reject_missing_reason(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)

            resp = client.post(
                f"/api/orders/{order.id}/reject",
                json={},
            )
            assert resp.status_code == 422  # validation error

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise


# ── Update Endpoint Tests ─────────────────────────────────────────────

class TestUpdateEndpoints:
    def test_update_order_header(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)

            resp = client.put(
                f"/api/orders/{order.id}",
                json={"customer_name": "Updated Customer", "po_number": "PO-UPDATED"},
            )
            assert resp.status_code == 200

            data = resp.json()
            assert data["customer_name"] == "Updated Customer"
            assert data["po_number"] == "PO-UPDATED"

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_update_order_not_pending(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)
            order.status = "approved"
            db.commit()

            resp = client.put(
                f"/api/orders/{order.id}",
                json={"customer_name": "Should Fail"},
            )
            assert resp.status_code == 400

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_update_line_item(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)
            item = order.line_items[0]

            resp = client.put(
                f"/api/orders/{order.id}/line-items/{item.id}",
                json={"quantity": 2000, "unit_price": "1.50"},
            )
            assert resp.status_code == 200

            data = resp.json()
            assert data["quantity"] == 2000
            assert float(data["unit_price"]) == 1.50
            assert float(data["line_total"]) == 3000.00  # 2000 * 1.50

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_update_line_item_not_found(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)
            fake_item_id = str(uuid.uuid4())

            resp = client.put(
                f"/api/orders/{order.id}/line-items/{fake_item_id}",
                json={"quantity": 100},
            )
            assert resp.status_code == 404

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise


# ── Form Download Endpoint Tests ──────────────────────────────────────

class TestFormDownloadEndpoints:
    def test_download_office_order(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)

            # Approve first to generate forms
            resp = client.post(f"/api/orders/{order.id}/approve")
            assert resp.status_code == 200

            # Download office order
            resp = client.get(f"/api/orders/{order.id}/forms/office-order")
            assert resp.status_code == 200
            assert "spreadsheetml" in resp.headers["content-type"]
            assert "Office_Order" in resp.headers["content-disposition"]
            assert len(resp.content) > 0

            # Verify it's a valid xlsx
            wb = load_workbook(BytesIO(resp.content))
            assert wb.active["A1"].value == "EEZY PEEZY PLASTICS"

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_download_office_order_before_generation(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)

            resp = client.get(f"/api/orders/{order.id}/forms/office-order")
            assert resp.status_code == 404
            assert "not yet generated" in resp.json()["detail"]

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_download_works_order(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)

            # Approve to generate forms
            resp = client.post(f"/api/orders/{order.id}/approve")
            assert resp.status_code == 200

            db.refresh(order)
            item = order.line_items[0]

            resp = client.get(f"/api/orders/{order.id}/forms/works-order/{item.id}")
            assert resp.status_code == 200
            assert "spreadsheetml" in resp.headers["content-type"]
            assert len(resp.content) > 0

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_download_works_order_before_generation(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)
            item = order.line_items[0]

            resp = client.get(f"/api/orders/{order.id}/forms/works-order/{item.id}")
            assert resp.status_code == 404

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_download_source_attachment(self, db):
        email = _create_test_email(db)
        try:
            # Add an attachment
            att = EmailAttachment(
                email_id=email.id,
                filename="purchase_order.pdf",
                content_type="application/pdf",
                file_data=b"%PDF-1.4 fake pdf content for testing",
                file_size_bytes=36,
            )
            db.add(att)
            db.commit()
            db.refresh(att)

            order = _create_test_order(db, email)

            resp = client.get(f"/api/orders/{order.id}/source-pdf/{att.id}")
            assert resp.status_code == 200
            assert resp.headers["content-type"] == "application/pdf"
            assert b"PDF" in resp.content

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_download_source_attachment_not_found(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)

            resp = client.get(f"/api/orders/{order.id}/source-pdf/999999")
            assert resp.status_code == 404

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise


# ── List/Detail with new fields ───────────────────────────────────────

class TestListDetailFields:
    def test_list_orders_has_forms_false(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)

            resp = client.get("/api/orders/")
            assert resp.status_code == 200
            data = resp.json()
            match = [o for o in data if o["id"] == str(order.id)]
            assert len(match) == 1
            assert match[0]["has_forms"] is False

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_list_orders_has_forms_true(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)

            # Approve to generate forms
            client.post(f"/api/orders/{order.id}/approve")

            resp = client.get("/api/orders/")
            data = resp.json()
            match = [o for o in data if o["id"] == str(order.id)]
            assert len(match) == 1
            assert match[0]["has_forms"] is True

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise

    def test_get_order_has_office_order(self, db):
        email = _create_test_email(db)
        try:
            order = _create_test_order(db, email)

            resp = client.get(f"/api/orders/{order.id}")
            assert resp.json()["has_office_order"] is False

            # Approve
            client.post(f"/api/orders/{order.id}/approve")

            resp = client.get(f"/api/orders/{order.id}")
            assert resp.json()["has_office_order"] is True

            _cleanup(db, [email.id])
        except Exception:
            _cleanup(db, [email.id])
            raise
