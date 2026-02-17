"""
Extraction service — sends email content (body + attachments) to Claude AI,
parses structured order data, matches product codes, and creates Order records.

All functions are synchronous (matching enrichment_service.py pattern).
Async wrapping is done at the API layer via run_in_executor.
"""

import base64
import json
import logging
import os
import re
import traceback
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Optional

import anthropic
import pdfplumber
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.models import IncomingEmail, Order, OrderLineItem
from app.services.enrichment_service import match_product_code

logger = logging.getLogger(__name__)


class ExtractionError(Exception):
    """Raised when LLM extraction fails."""
    pass


# ── System prompt for Claude ─────────────────────────────────────────

EXTRACTION_SYSTEM_PROMPT = """You are an order extraction assistant for Ramjet Plastics, a plastics manufacturing company.

IMPORTANT: Ramjet Plastics is the SELLER/MANUFACTURER. The customer is the company or person PLACING the order (i.e. the sender of the email). Never set customer_name to "Ramjet Plastics" — that is us. The customer_name should be the buying company or person, which can typically be identified from the email sender, the "Bill To" / "Ship To" fields, or the company placing the purchase order.

Your job is to extract structured purchase order data from emails and their attachments (PDFs, Excel files, images).

Return a JSON object with this exact structure:

{
  "customer_name": { "value": "<string>", "confidence": <0.0-1.0> },
  "po_number": { "value": "<string>", "confidence": <0.0-1.0> },
  "po_date": { "value": "<YYYY-MM-DD or null>", "confidence": <0.0-1.0> },
  "delivery_date": { "value": "<YYYY-MM-DD or null>", "confidence": <0.0-1.0> },
  "special_instructions": { "value": "<string or null>", "confidence": <0.0-1.0> },
  "line_items": [
    {
      "product_code": { "value": "<string>", "confidence": <0.0-1.0> },
      "description": { "value": "<string>", "confidence": <0.0-1.0> },
      "quantity": { "value": <integer>, "confidence": <0.0-1.0> },
      "colour": { "value": "<string or null>", "confidence": <0.0-1.0> },
      "unit_price": { "value": <number or null>, "confidence": <0.0-1.0> }
    }
  ],
  "overall_confidence": <0.0-1.0>
}

Rules:
- Extract ALL line items from the order
- Product codes are typically short alphanumeric strings (e.g. "LOCAP2", "GLCAPRB")
- Dates should be in YYYY-MM-DD format
- If a field is not found, set value to null and confidence to 0.0
- Confidence should reflect how certain you are about the extraction
- Only return the JSON object, no other text
- If the email does not appear to contain a purchase order, return:
  {"error": "No purchase order found", "overall_confidence": 0.0}
"""


# ── Content preparation ──────────────────────────────────────────────

def prepare_email_content(email: IncomingEmail) -> list[dict]:
    """
    Build Claude API content blocks from email body + attachments.
    Returns a list of content blocks suitable for the messages API.
    """
    content_blocks = []

    # Email body text
    body = email.body_text or ""
    if body.strip():
        content_blocks.append({
            "type": "text",
            "text": f"=== EMAIL BODY ===\nFrom: {email.sender or 'Unknown'}\nSubject: {email.subject or 'No subject'}\n\n{body}"
        })

    # Process attachments
    for attachment in (email.attachments or []):
        if not attachment.file_data:
            continue

        content_type = (attachment.content_type or "").lower()
        filename = attachment.filename or "unknown"

        if content_type == "application/pdf" or filename.lower().endswith(".pdf"):
            block = _prepare_pdf_block(attachment.file_data, filename)
            if block:
                content_blocks.append(block)

        elif content_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ) or filename.lower().endswith((".xlsx", ".xls")):
            text = parse_excel_to_text(attachment.file_data, filename)
            if text:
                content_blocks.append({
                    "type": "text",
                    "text": f"=== EXCEL ATTACHMENT: {filename} ===\n{text}"
                })

        elif content_type.startswith("image/"):
            media_type = content_type
            # Ensure valid media type for Claude
            if media_type not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
                media_type = "image/png"
            content_blocks.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": media_type,
                    "data": base64.standard_b64encode(attachment.file_data).decode("utf-8"),
                }
            })

    # If no content was prepared, add a fallback
    if not content_blocks:
        content_blocks.append({
            "type": "text",
            "text": f"Email from {email.sender or 'Unknown'}, subject: {email.subject or 'No subject'}. No extractable content found."
        })

    return content_blocks


def _prepare_pdf_block(file_data: bytes, filename: str) -> Optional[dict]:
    """
    Prepare a PDF content block. Uses Claude's native PDF support (base64 document).
    Falls back to text extraction for oversized PDFs (>25MB).
    """
    MAX_PDF_SIZE = 25 * 1024 * 1024  # 25MB limit for Claude API

    if len(file_data) <= MAX_PDF_SIZE:
        return {
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": base64.standard_b64encode(file_data).decode("utf-8"),
            }
        }

    # Fallback: extract text from oversized PDFs
    logger.warning(f"PDF {filename} exceeds 25MB, falling back to text extraction")
    text = extract_pdf_text_fallback(file_data)
    if text:
        return {
            "type": "text",
            "text": f"=== PDF ATTACHMENT (text extracted): {filename} ===\n{text}"
        }
    return None


def parse_excel_to_text(file_data: bytes, filename: str) -> str:
    """Convert Excel bytes to readable structured text for LLM consumption."""
    try:
        wb = load_workbook(filename=BytesIO(file_data), data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"--- Sheet: {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                cell_values = [str(c) if c is not None else "" for c in row]
                if any(v.strip() for v in cell_values):
                    lines.append("\t".join(cell_values))
        return "\n".join(lines)
    except Exception as e:
        logger.error(f"Failed to parse Excel file {filename}: {e}")
        return ""


def extract_pdf_text_fallback(file_data: bytes) -> str:
    """Extract text from PDF using pdfplumber (fallback for oversized PDFs)."""
    try:
        with pdfplumber.open(BytesIO(file_data)) as pdf:
            pages = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages)
    except Exception as e:
        logger.error(f"Failed to extract PDF text: {e}")
        return ""


# ── LLM extraction ───────────────────────────────────────────────────

def extract_order_from_email(email: IncomingEmail) -> dict:
    """
    Send email content to Claude API and parse the structured JSON response.
    Returns the parsed extraction dict.
    Raises ExtractionError on failure.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise ExtractionError("ANTHROPIC_API_KEY not configured")

    model = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-5-20250929")
    content_blocks = prepare_email_content(email)

    client = anthropic.Anthropic(api_key=api_key)

    try:
        response = client.messages.create(
            model=model,
            max_tokens=4096,
            system=EXTRACTION_SYSTEM_PROMPT,
            messages=[
                {"role": "user", "content": content_blocks}
            ],
        )
    except anthropic.APIError as e:
        raise ExtractionError(f"Anthropic API error: {e}")

    # Extract text from response
    response_text = ""
    for block in response.content:
        if block.type == "text":
            response_text += block.text

    if not response_text.strip():
        raise ExtractionError("Empty response from Claude API")

    # Parse JSON from response
    return _parse_extraction_json(response_text)


def _parse_extraction_json(text: str) -> dict:
    """
    Parse JSON from Claude's response text.
    Handles both clean JSON and responses with surrounding text.
    """
    # Try direct parse first
    try:
        return json.loads(text.strip())
    except json.JSONDecodeError:
        pass

    # Try to find JSON block in markdown code fence
    code_block_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, re.DOTALL)
    if code_block_match:
        try:
            return json.loads(code_block_match.group(1))
        except json.JSONDecodeError:
            pass

    # Fallback: find the outermost { ... } block
    brace_match = re.search(r'\{.*\}', text, re.DOTALL)
    if brace_match:
        try:
            return json.loads(brace_match.group(0))
        except json.JSONDecodeError:
            pass

    raise ExtractionError(f"Could not parse JSON from response: {text[:200]}...")


# ── Order creation ────────────────────────────────────────────────────

def create_order_from_extraction(db: Session, email: IncomingEmail, extraction: dict) -> Optional[Order]:
    """
    Create Order + OrderLineItem records from extraction data.
    Matches product codes against the catalog and flags items needing review.
    Returns None if an order already exists for this email (duplicate guard).
    """
    # Guard against duplicate orders for the same email
    existing = db.query(Order).filter(Order.email_id == email.id).first()
    if existing:
        logger.warning(f"Order {existing.id} already exists for email {email.id} — skipping duplicate")
        email.processed = True
        db.commit()
        return existing

    # Check for extraction error
    if "error" in extraction:
        order = Order(
            email_id=email.id,
            status="error",
            extraction_raw_json=extraction,
            extraction_confidence=Decimal("0.00"),
        )
        db.add(order)
        email.processed = True
        db.commit()
        return order

    # Extract header fields
    def _get_field(field_name: str, default=None):
        field = extraction.get(field_name, {})
        if isinstance(field, dict):
            return field.get("value", default)
        return field if field is not None else default

    def _get_confidence(field_name: str) -> Decimal:
        field = extraction.get(field_name, {})
        if isinstance(field, dict):
            try:
                return Decimal(str(field.get("confidence", 0)))
            except (InvalidOperation, TypeError):
                return Decimal("0.00")
        return Decimal("0.00")

    # Parse po_date
    po_date = None
    po_date_str = _get_field("po_date")
    if po_date_str:
        try:
            from datetime import date
            po_date = date.fromisoformat(po_date_str)
        except (ValueError, TypeError):
            pass

    # Parse delivery_date
    delivery_date = None
    delivery_date_str = _get_field("delivery_date")
    if delivery_date_str:
        try:
            from datetime import date
            delivery_date = date.fromisoformat(delivery_date_str)
        except (ValueError, TypeError):
            pass

    overall_confidence = extraction.get("overall_confidence", 0)
    try:
        overall_confidence = Decimal(str(overall_confidence))
    except (InvalidOperation, TypeError):
        overall_confidence = Decimal("0.00")

    order = Order(
        email_id=email.id,
        status="pending",
        customer_name=_get_field("customer_name"),
        po_number=_get_field("po_number"),
        po_date=po_date,
        delivery_date=delivery_date,
        special_instructions=_get_field("special_instructions"),
        extraction_confidence=overall_confidence,
        extraction_raw_json=extraction,
    )
    db.add(order)
    db.flush()  # Get the order ID

    # Process line items
    line_items_data = extraction.get("line_items", [])
    for idx, item_data in enumerate(line_items_data, start=1):
        def _item_field(name, default=None):
            field = item_data.get(name, {})
            if isinstance(field, dict):
                return field.get("value", default)
            return field if field is not None else default

        def _item_confidence(name) -> Decimal:
            field = item_data.get(name, {})
            if isinstance(field, dict):
                try:
                    return Decimal(str(field.get("confidence", 0)))
                except (InvalidOperation, TypeError):
                    return Decimal("0.00")
            return Decimal("0.00")

        extracted_code = _item_field("product_code", "")
        quantity = _item_field("quantity", 0)
        unit_price_val = _item_field("unit_price")

        # Convert types
        try:
            quantity = int(quantity) if quantity else 0
        except (ValueError, TypeError):
            quantity = 0

        unit_price = None
        if unit_price_val is not None:
            try:
                unit_price = Decimal(str(unit_price_val))
            except (InvalidOperation, TypeError):
                unit_price = None

        line_total = None
        if unit_price is not None and quantity:
            line_total = (unit_price * quantity).quantize(Decimal("0.01"))

        # Calculate average item confidence
        item_confidences = [
            _item_confidence("product_code"),
            _item_confidence("quantity"),
            _item_confidence("description"),
        ]
        avg_confidence = sum(item_confidences) / len(item_confidences) if item_confidences else Decimal("0.00")

        # Match product code against catalog
        matched_code = None
        needs_review = False

        if extracted_code:
            matches = match_product_code(db, extracted_code)
            if matches:
                best_match = matches[0]
                match_confidence = best_match.get("confidence", 0)
                if match_confidence >= 0.9:
                    matched_code = best_match["product_code"]
                else:
                    matched_code = best_match["product_code"]
                    needs_review = True
            else:
                needs_review = True
        else:
            needs_review = True

        # Also flag if item confidence is low
        if avg_confidence < Decimal("0.80"):
            needs_review = True

        line_item = OrderLineItem(
            order_id=order.id,
            line_number=idx,
            product_code=extracted_code,
            matched_product_code=matched_code,
            product_description=_item_field("description"),
            colour=_item_field("colour"),
            quantity=quantity,
            unit_price=unit_price,
            line_total=line_total,
            confidence=avg_confidence,
            needs_review=needs_review,
        )
        db.add(line_item)

    # Mark email as processed
    email.processed = True
    db.commit()
    db.refresh(order)

    return order


# ── Pipeline orchestration ────────────────────────────────────────────

def process_single_email(db: Session, email: IncomingEmail) -> Optional[Order]:
    """
    Full pipeline for one email: extract → match → create order.
    Returns the created Order, or None on failure.
    """
    # Skip if already processed (race condition guard)
    if email.processed:
        existing = db.query(Order).filter(Order.email_id == email.id).first()
        if existing:
            logger.info(f"Email {email.id} already processed — order {existing.id} exists")
            return existing

    try:
        logger.info(f"Processing email {email.id}: '{email.subject}'")
        extraction = extract_order_from_email(email)
        order = create_order_from_extraction(db, email, extraction)
        logger.info(f"Created order {order.id} from email {email.id} (status={order.status})")
        return order
    except ExtractionError as e:
        logger.error(f"Extraction failed for email {email.id}: {e}")
        logger.debug(traceback.format_exc())
        return _create_error_order(db, email, str(e))
    except Exception as e:
        logger.error(f"Unexpected error processing email {email.id}: {e}")
        logger.debug(traceback.format_exc())
        return _create_error_order(db, email, str(e))


def _create_error_order(db: Session, email: IncomingEmail, error_msg: str) -> Order:
    """Create an error-status order for visibility when extraction fails."""
    existing = db.query(Order).filter(Order.email_id == email.id).first()
    if existing:
        logger.warning(f"Order {existing.id} already exists for email {email.id} — skipping error order")
        return existing

    order = Order(
        email_id=email.id,
        status="error",
        customer_name=email.sender,
        extraction_confidence=Decimal("0.00"),
        extraction_raw_json={"error": error_msg},
    )
    db.add(order)
    email.processed = True
    db.commit()
    db.refresh(order)
    return order


def process_unprocessed_emails(db: Session) -> dict:
    """
    Process all unprocessed emails.
    Returns summary dict with orders_created and errors counts.
    """
    unprocessed = (
        db.query(IncomingEmail)
        .filter(IncomingEmail.processed == False)
        .order_by(IncomingEmail.received_at)
        .all()
    )

    if not unprocessed:
        return {"orders_created": 0, "errors": 0, "message": "No unprocessed emails found"}

    orders_created = 0
    errors = 0

    for email in unprocessed:
        order = process_single_email(db, email)
        if order and order.status != "error":
            orders_created += 1
        else:
            errors += 1

    return {
        "orders_created": orders_created,
        "errors": errors,
        "message": f"Processed {len(unprocessed)} emails: {orders_created} orders created, {errors} errors",
    }
