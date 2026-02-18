"""
Form generation service — builds Office Order and Works Order Excel files
from enriched order data using openpyxl.
"""

import logging
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.models import Order, OrderLineItem
from app.services.enrichment_service import (
    calculate_material_requirements,
    get_product_full_specs,
)

logger = logging.getLogger(__name__)

# ── Shared style constants ────────────────────────────────────────────

_TITLE_FONT = Font(name="Calibri", size=16, bold=True)
_SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True)
_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_LABEL_FONT = Font(name="Calibri", size=10, bold=True)
_BODY_FONT = Font(name="Calibri", size=10)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _to_float(val) -> Optional[float]:
    """Convert Decimal / numeric to float for openpyxl, or return None."""
    if val is None:
        return None
    if isinstance(val, Decimal):
        return float(val)
    return float(val)


def _fmt_date(d) -> str:
    """Format a date as dd/mm/YYYY (Australian convention)."""
    if d is None:
        return ""
    if isinstance(d, date):
        return d.strftime("%d/%m/%Y")
    return str(d)


# ── Office Order ──────────────────────────────────────────────────────

def generate_office_order(db: Session, order: Order) -> bytes:
    """Build an Office Order .xlsx for one order and return the bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Office Order"

    # Column widths (A-G)
    col_widths = [8, 16, 35, 14, 10, 14, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1 — Company title
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "RAMJET PLASTICS"
    cell.font = _TITLE_FONT
    cell.alignment = Alignment(horizontal="center")

    # Row 2 — Form title
    ws.merge_cells("A2:G2")
    cell = ws["A2"]
    cell.value = "OFFICE ORDER"
    cell.font = _SUBTITLE_FONT
    cell.alignment = Alignment(horizontal="center")

    # Row 4 — Customer / PO#
    ws["A4"].value = "Customer:"
    ws["A4"].font = _LABEL_FONT
    ws.merge_cells("B4:C4")
    ws["B4"].value = order.customer_name or ""
    ws["B4"].font = _BODY_FONT

    ws["E4"].value = "PO #:"
    ws["E4"].font = _LABEL_FONT
    ws.merge_cells("F4:G4")
    ws["F4"].value = order.po_number or ""
    ws["F4"].font = _BODY_FONT

    # Row 5 — Dates
    ws["A5"].value = "PO Date:"
    ws["A5"].font = _LABEL_FONT
    ws.merge_cells("B5:C5")
    ws["B5"].value = _fmt_date(order.po_date)
    ws["B5"].font = _BODY_FONT

    ws["E5"].value = "Delivery:"
    ws["E5"].font = _LABEL_FONT
    ws.merge_cells("F5:G5")
    ws["F5"].value = _fmt_date(order.delivery_date)
    ws["F5"].font = _BODY_FONT

    # Row 7 — Table headers
    headers = ["Line", "Product Code", "Description", "Colour", "Qty", "Unit Price", "Line Total"]
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=7, column=col_idx, value=header)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.border = _THIN_BORDER
        c.alignment = Alignment(horizontal="center")

    # Row 8+ — Line items
    items = sorted(order.line_items, key=lambda li: li.line_number)
    row = 8
    order_total = Decimal("0.00")
    for idx, item in enumerate(items):
        values = [
            item.line_number,
            item.product_code or item.matched_product_code or "",
            item.product_description or "",
            item.colour or "",
            item.quantity,
            _to_float(item.unit_price),
            _to_float(item.line_total),
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = _BODY_FONT
            c.border = _THIN_BORDER
            if idx % 2 == 1:
                c.fill = _ALT_ROW_FILL
            # Right-align numeric columns
            if col_idx in (5, 6, 7):
                c.alignment = Alignment(horizontal="right")
            # Currency format
            if col_idx in (6, 7) and val is not None:
                c.number_format = '#,##0.00'

        if item.line_total:
            order_total += Decimal(str(item.line_total)).quantize(Decimal("0.01"))
        row += 1

    # Subtotal row
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="Subtotal")
    c.font = _LABEL_FONT
    c.fill = _TOTAL_FILL
    c.alignment = Alignment(horizontal="right")
    c.border = _THIN_BORDER
    for col_idx in range(2, 7):
        mc = ws.cell(row=row, column=col_idx)
        mc.fill = _TOTAL_FILL
        mc.border = _THIN_BORDER

    subtotal_cell = ws.cell(row=row, column=7, value=_to_float(order_total))
    subtotal_cell.font = _LABEL_FONT
    subtotal_cell.fill = _TOTAL_FILL
    subtotal_cell.border = _THIN_BORDER
    subtotal_cell.number_format = '#,##0.00'
    subtotal_cell.alignment = Alignment(horizontal="right")

    # GST row
    gst_amount = (order_total * Decimal("0.10")).quantize(Decimal("0.01"))
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="GST (10%)")
    c.font = _LABEL_FONT
    c.fill = _TOTAL_FILL
    c.alignment = Alignment(horizontal="right")
    c.border = _THIN_BORDER
    for col_idx in range(2, 7):
        mc = ws.cell(row=row, column=col_idx)
        mc.fill = _TOTAL_FILL
        mc.border = _THIN_BORDER

    gst_cell = ws.cell(row=row, column=7, value=_to_float(gst_amount))
    gst_cell.font = _LABEL_FONT
    gst_cell.fill = _TOTAL_FILL
    gst_cell.border = _THIN_BORDER
    gst_cell.number_format = '#,##0.00'
    gst_cell.alignment = Alignment(horizontal="right")

    # Total (inc. GST) row
    grand_total = (order_total + gst_amount).quantize(Decimal("0.01"))
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="Total (inc. GST)")
    c.font = Font(name="Calibri", size=10, bold=True)
    c.fill = _TOTAL_FILL
    c.alignment = Alignment(horizontal="right")
    c.border = _THIN_BORDER
    for col_idx in range(2, 7):
        mc = ws.cell(row=row, column=col_idx)
        mc.fill = _TOTAL_FILL
        mc.border = _THIN_BORDER

    grand_cell = ws.cell(row=row, column=7, value=_to_float(grand_total))
    grand_cell.font = Font(name="Calibri", size=10, bold=True)
    grand_cell.fill = _TOTAL_FILL
    grand_cell.border = _THIN_BORDER
    grand_cell.number_format = '#,##0.00'
    grand_cell.alignment = Alignment(horizontal="right")

    # Special Instructions
    if order.special_instructions:
        row += 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value="Special Instructions")
        c.font = _LABEL_FONT

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=order.special_instructions)
        c.font = _BODY_FONT
        c.alignment = Alignment(wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Works Order ───────────────────────────────────────────────────────

def _write_section_header(ws, row: int, title: str, end_col: int = 6) -> int:
    """Write a blue section header spanning columns A to end_col. Returns next row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=1, value=title)
    c.font = _SECTION_FONT
    c.fill = _SECTION_FILL
    c.border = _THIN_BORDER
    for col_idx in range(2, end_col + 1):
        mc = ws.cell(row=row, column=col_idx)
        mc.fill = _SECTION_FILL
        mc.border = _THIN_BORDER
    return row + 1


def _write_kv_row(ws, row: int, label: str, value, label_col: int = 1, value_col: int = 2, merge_to: int = 3) -> int:
    """Write a label-value pair. Returns next row."""
    c = ws.cell(row=row, column=label_col, value=label)
    c.font = _LABEL_FONT
    c.border = _THIN_BORDER

    if merge_to > value_col:
        ws.merge_cells(start_row=row, start_column=value_col, end_row=row, end_column=merge_to)

    display = value
    if isinstance(value, Decimal):
        display = float(value)
    if isinstance(value, date):
        display = _fmt_date(value)

    vc = ws.cell(row=row, column=value_col, value=display)
    vc.font = _BODY_FONT
    vc.border = _THIN_BORDER

    return row + 1


def _write_kv_pair_row(ws, row: int, label1: str, val1, label2: str, val2) -> int:
    """Write two label-value pairs side by side (cols A-C + D-F). Returns next row."""
    row = _write_kv_row(ws, row, label1, val1, label_col=1, value_col=2, merge_to=3)
    # Go back to the same row for the second pair
    actual_row = row - 1
    c = ws.cell(row=actual_row, column=4, value=label2)
    c.font = _LABEL_FONT
    c.border = _THIN_BORDER

    display = val2
    if isinstance(val2, Decimal):
        display = float(val2)
    if isinstance(val2, date):
        display = _fmt_date(val2)

    ws.merge_cells(start_row=actual_row, start_column=5, end_row=actual_row, end_column=6)
    vc = ws.cell(row=actual_row, column=5, value=display)
    vc.font = _BODY_FONT
    vc.border = _THIN_BORDER

    return row


def generate_works_order(
    db: Session,
    order: Order,
    line_item: OrderLineItem,
    adjusted_quantity: Optional[int] = None,
    verified_stock: Optional[int] = None,
) -> bytes:
    """
    Build a Works Order .xlsx for one line item and return the bytes.

    If adjusted_quantity is provided, uses it instead of line_item.quantity
    for the production quantity. Adds a stock note if verified_stock is given.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Works Order"

    # The quantity to display on the WO
    production_qty = adjusted_quantity if adjusted_quantity is not None else line_item.quantity

    # Column widths (A-F)
    col_widths = [20, 20, 20, 20, 20, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wo_number = f"WO-{order.po_number or 'N/A'}-{line_item.line_number}"

    # Title rows
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "RAMJET PLASTICS"
    c.font = _TITLE_FONT
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "WORKS ORDER"
    c.font = _SUBTITLE_FONT
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:F3")
    c = ws["A3"]
    c.value = wo_number
    c.font = _SUBTITLE_FONT
    c.alignment = Alignment(horizontal="center")

    row = 5

    # ── Section 1: Order Details ──────────────────────────────────────
    row = _write_section_header(ws, row, "Order Details")
    row = _write_kv_pair_row(ws, row, "WO #:", wo_number, "PO #:", order.po_number or "")
    row = _write_kv_pair_row(ws, row, "Customer:", order.customer_name or "", "PO Date:", order.po_date)
    row = _write_kv_pair_row(ws, row, "Product Code:", line_item.product_code or line_item.matched_product_code or "", "Delivery Date:", order.delivery_date)
    row = _write_kv_pair_row(ws, row, "Description:", line_item.product_description or "", "Colour:", line_item.colour or "")
    row = _write_kv_pair_row(ws, row, "Ordered Qty:", line_item.quantity, "Produce Qty:", production_qty)

    # Stock note if verified stock was deducted
    if verified_stock is not None and verified_stock > 0:
        stock_note = f"Stock on hand: {verified_stock} — Produce: {production_qty}"
        row = _write_kv_row(ws, row, "Stock Note:", stock_note, merge_to=6)

    row = _write_kv_pair_row(ws, row, "Line Total:", _to_float(line_item.line_total), "", "")
    row += 1

    # Fetch product specs if matched
    specs = None
    mat_reqs = None
    product_code = line_item.matched_product_code
    colour = line_item.colour

    if product_code:
        specs = get_product_full_specs(db, product_code, colour)
        # get_product_full_specs returns a single dict when colour is specified, a list otherwise
        if isinstance(specs, list):
            specs = specs[0] if specs else None

        if colour:
            mat_reqs = calculate_material_requirements(db, product_code, colour, production_qty)

    # ── Section 2: Manufacturing Specifications ───────────────────────
    row = _write_section_header(ws, row, "Manufacturing Specifications")
    if specs:
        row = _write_kv_pair_row(ws, row, "Mould No:", specs.get("mould_no", ""), "Num Cavities:", specs.get("num_cavities", ""))
        row = _write_kv_pair_row(ws, row, "Cycle Time (s):", specs.get("cycle_time_seconds", ""), "Shot Weight (g):", specs.get("shot_weight_grams", ""))
        row = _write_kv_pair_row(ws, row, "Product Weight (g):", specs.get("product_weight_grams", ""), "Est. Running Time (hrs):", specs.get("estimated_running_time_hours", ""))
        row = _write_kv_row(ws, row, "Machine Min Req:", specs.get("machine_min_requirements", ""), merge_to=6)
    else:
        row = _write_kv_row(ws, row, "No matched product — specifications unavailable", "", merge_to=6)
    row += 1

    # ── Section 3: Material Specifications ────────────────────────────
    row = _write_section_header(ws, row, "Material Specifications")
    if specs and specs.get("material"):
        mat = specs["material"]
        row = _write_kv_pair_row(ws, row, "Material Grade:", mat.get("material_grade", ""), "Material Type:", mat.get("material_type", ""))
        row = _write_kv_pair_row(ws, row, "Colour:", mat.get("colour", ""), "Colour No:", mat.get("colour_no", ""))
        row = _write_kv_pair_row(ws, row, "Colour Supplier:", mat.get("colour_supplier", ""), "MB Add Rate (%):", mat.get("mb_add_rate", ""))
        row = _write_kv_pair_row(ws, row, "Additive:", mat.get("additive", ""), "Additive Add Rate (%):", mat.get("additive_add_rate", ""))
        row = _write_kv_row(ws, row, "Additive Supplier:", mat.get("additive_supplier", ""), merge_to=6)
    else:
        row = _write_kv_row(ws, row, "No matched product — material specs unavailable", "", merge_to=6)
    row += 1

    # ── Section 4: Material Requirements (Calculated) ─────────────────
    row = _write_section_header(ws, row, "Material Requirements (Calculated)")
    if mat_reqs and mat_reqs.get("material_requirements"):
        mr = mat_reqs["material_requirements"]
        row = _write_kv_pair_row(ws, row, "Base Material (kg):", mr.get("base_material_kg", ""), "Material Grade:", mr.get("material_grade", ""))
        row = _write_kv_pair_row(ws, row, "Masterbatch (kg):", mr.get("masterbatch_kg", ""), "Colour No:", mr.get("colour_no", ""))
        row = _write_kv_pair_row(ws, row, "Additive (kg):", mr.get("additive_kg", ""), "Additive:", mr.get("additive", ""))
        row = _write_kv_row(ws, row, "Total Material (kg):", mr.get("total_material_kg", ""), merge_to=6)
    else:
        row = _write_kv_row(ws, row, "No matched product — cannot calculate requirements", "", merge_to=6)
    row += 1

    # ── Section 5: Packaging ──────────────────────────────────────────
    row = _write_section_header(ws, row, "Packaging")
    if specs and specs.get("packaging"):
        pkg = specs["packaging"]
        row = _write_kv_pair_row(ws, row, "Qty per Bag:", pkg.get("qty_per_bag", ""), "Bag Size:", pkg.get("bag_size", ""))
        row = _write_kv_pair_row(ws, row, "Qty per Carton:", pkg.get("qty_per_carton", ""), "Carton Size:", pkg.get("carton_size", ""))
        row = _write_kv_pair_row(ws, row, "Cartons per Pallet:", pkg.get("cartons_per_pallet", ""), "Cartons per Layer:", pkg.get("cartons_per_layer", ""))
    else:
        row = _write_kv_row(ws, row, "No matched product — packaging specs unavailable", "", merge_to=6)
    row += 1

    # ── Section 6: Packaging Requirements (Calculated) ────────────────
    row = _write_section_header(ws, row, "Packaging Requirements (Calculated)")
    if mat_reqs and mat_reqs.get("packaging_requirements"):
        pr = mat_reqs["packaging_requirements"]
        row = _write_kv_pair_row(ws, row, "Bags Needed:", pr.get("bags_needed", ""), "Bag Size:", pr.get("bag_size", ""))
        row = _write_kv_pair_row(ws, row, "Cartons Needed:", pr.get("cartons_needed", ""), "Carton Size:", pr.get("carton_size", ""))
    else:
        row = _write_kv_row(ws, row, "No matched product — cannot calculate requirements", "", merge_to=6)
    row += 1

    # ── Section 7: Notes ──────────────────────────────────────────────
    row = _write_section_header(ws, row, "Notes")
    if order.special_instructions:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=order.special_instructions)
        c.font = _BODY_FONT
        c.alignment = Alignment(wrap_text=True)
        c.border = _THIN_BORDER
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value="")
        c.border = _THIN_BORDER

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Orchestrator ──────────────────────────────────────────────────────

def generate_all_forms(db: Session, order: Order) -> None:
    """
    Generate Office Order + one Works Order per line item.
    Stores the bytes on the model fields and commits.
    """
    logger.info("Generating forms for order %s (PO: %s)", order.id, order.po_number)

    # Office Order
    office_bytes = generate_office_order(db, order)
    order.office_order_file = office_bytes
    logger.info("Office order generated: %d bytes", len(office_bytes))

    # Works Orders — one per line item
    for item in order.line_items:
        wo_bytes = generate_works_order(db, order, item)
        item.works_order_file = wo_bytes
        logger.info(
            "Works order WO-%s-%d generated: %d bytes",
            order.po_number, item.line_number, len(wo_bytes),
        )

    db.commit()
    logger.info("All forms generated and saved for order %s", order.id)
